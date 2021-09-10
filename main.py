import mysql.connector
import pandas as pd

mydb = mysql.connector.connect(
host = 'localhost',
user='root',
password='warda170199',
port=3306,
database='coursedb'
)

exl_to_csv= pd.read_csv (r'C:\Users\warda.kashif\PycharmProjects\DataHandling\Student.csv.txt')
exl_to_csv.to_excel ('Student.xlsx')

cur=pd.read_sql('Select * from course', con = mydb)

df_sql_data = pd.DataFrame(cur)
df_sql_data.head()
df_sql_data.to_excel("output.xlsx")

pd.set_option('display.expand_frame_repr', False,)


mydb.close()
# Data Coming From Database
print("Data Coming From Database")
df1=cur
print(df1)


print("\n\nData Change from csv to excel\n")
df2=pd.read_excel("Student.xlsx",index_col=0)
print(df2)

print("\n\nData which is already in excel format\n")
df3=pd.read_excel("Record.xlsx")
print(df3)


df4 = df2.merge(df3, on="STUDENTID", how='outer')
print(df4)

df5=df4.merge(df1, on="courseId", how='outer')
print(df5)

df5.to_excel("final_output.xlsx")


#Second Task
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

wb = load_workbook('sheet.xlsx')
sheet = wb['Sheet1']
# cell references (original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column)
print(max_column)
print(min_row)
print(max_row)


sheet['G2'] = '=(C2/100)*100'


import string
alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column]
print(excel_alphabet)


for i in excel_alphabet:
    if i=='G':
        for j in range(max_row):
            sheet[f'{i}{j + 2}']= f'=(C{j + 2}/100 * 100)'

wb.save('sheet1.xlsx')


# adding total label
# sheet[f'{excel_alphabet[0]}{max_column+1}'] = 'Total'
# wb.save('sheet1.xlsx')







import pandas as pd

df=pd.read_excel("final_output.xlsx",index_col=0)
print(df)

per=[]
for val in df["Marks"]:
    percent=((val)/100) * 100
    per.append(percent)
print(per)

df["Percentage"]=per
print(df)
total=50
# df.style.apply(lambda x:["background:red" if x<total else "background:green" for x in df.Percentage], axis=0)
df2=df.style.apply(lambda x:["background-color:red" if x<total else "background-color:green" for x in df.Percentage], axis=0)


df2.to_excel("sheet3.xlsx", engine="openpyxl" ,index=False)


