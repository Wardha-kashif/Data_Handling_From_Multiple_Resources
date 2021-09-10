import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting import formatting
from openpyxl.styles.builtins import styles
from openpyxl.utils import get_column_letter

df=pd.read_excel('sheet4.xlsx',index_col=0)
per=[]
for val in df["Marks"]:
    percent=((val)/100) * 100
    per.append(percent)
print(per)

df["Percentage"]=per
print(df)

df.to_excel("sheet4.xlsx")

df=pd.read_excel("sheet4.xlsx",index_col=0)
print(df)
total=50
# df.style.apply(lambda x:["background:red" if x<total else "background:green" for x in df.Percentage], axis=0)
df2=df.style.apply(lambda x:["background-color:red" if x<total else "background-color:green" for x in df.Percentage], axis=0)

df2.to_excel("sheet.xlsx", engine="openpyxl" ,index=False)





























# min_column = wb.active.min_column
# max_column = wb.active.max_column
# min_row = wb.active.min_row
# max_row = wb.active.max_row


# for row in range(1,7):
#     for column in range(1,7):
#         ws.cell(row=row, column=column)
#         ws.conditional_formatting.add('H',formatting.rule.CellIsRule(operator='lessThan', formula=['50'],fill=red_fill))
#
# wb.save('sheet4.xlsx')
# #
# import pyexcel as pe
# records = pe.iget_records(file_name="final_output.xlsx")
# #import pyexcel
# # records = pyexcel.get_book(file_name="final_output.xlsx")
#
# for record in records:
#     obt_Marks = record["Marks"]
#     quotient = obt_Marks / 100
#     percent = int(quotient * 100)
#     print(percent)
# .conditional_format('Marks', {'type':     'cell',
#                                        'criteria': 'between',
#                                        'maximum':  30,
#                                        'format':  green_format})
#
# # path = "C:\\Users\\warda.kashif\\PycharmProjects\\DataHandling\\final_output.xlsx"
# #
# # wb_obj = openpyxl.load_workbook(path)
# #
# # sheet_obj = wb_obj.active
# # df=sheet_obj.to_excel()
# # for rows in wb_obj (min_row=1, max_row=1, min_col=1):
# #     for cell in rows:
# #       cell.fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
#
# # Python program to read an excel file
#
# # # import openpyxl module
# # import openpyxl
# #
# # # Give the location of the file
# # path = "C:\\Users\\warda.kashif\\PycharmProjects\\DataHandling\\final_output.xlsx"
# #
# # wb_obj = openpyxl.load_workbook(path)
# #
# # sheet_obj = wb_obj.active
# #
# # # for record in wb_obj:
# # #     obt_Marks = record["Marks"]
# # #     quotient = obt_Marks / 100
# # #     percent = int(quotient * 100)
# # #     print(percent)
# # max_col = sheet_obj.max_column
# # for i in range(1, max_col + 1):
# #     cell_obj = sheet_obj.cell(row = 2, column = i)
# #     print(cell_obj.value, end = " ")
#
#
# # from openpyxl import load_workbook
# # wb = load_workbook(filename = 'final_output.xlsx')
# # sheet_ranges = wb['Sheet1']
# # print(sheet_ranges)
